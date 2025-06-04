import openpyxl
import re
import pandas as pd

def hallar_precedentes_de_hoja(ruta_acceso_documento=""):
    wb = openpyxl.load_workbook(ruta_acceso_documento, data_only=False)
    ws = wb.active 
    
    precedentes = {}
    ref_regex = re.compile(r"\b([A-Z]{1,3}\d+)\b")

    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula = cell.value
                refs = ref_regex.findall(formula)
                precedentes[cell.coordinate] = refs
    
    return precedentes


def leer_nombres_hojas(ruta_acceso_documento=""):
    archivo = pd.ExcelFile(ruta_acceso_documento)
    return archivo.sheet_names

def hallar_precedentes_hoja_externa(ruta_acceso_documento=""):
    wb = openpyxl.load_workbook(ruta_acceso_documento)
    ref_regex = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_]+))!([A-Z]{1,3}\d+)")

    precedentes = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and isinstance(cell.value  , str):
                    formula = cell.value

                    matches = ref_regex.findall(formula)
                    refs = []

                    for match in matches:
                        hoja = match[0] if match[0] else match[1]
                        celda = match[2]

                        refs.append(f"{hoja}!{celda}")
                    
                    if refs :
                        precedentes[f"{sheet}!{cell.coordinate}"] = refs
    return precedentes

def leer_precedentes_hoja_externa(data_precedentes={}):
    for celda, refs in data_precedentes.items():
        print(f"{celda} depende de: {','.join(refs)}")

def leer_precedentes_hoja(data_precedentes={}):
    for celda, refs in data_precedentes.items():
        print(f"{celda} depende de: {', '.join(refs)}")

def obtener_precedentes_por_hoja(ruta_excel, nombre_hoja):
    wb = openpyxl.load_workbook(ruta_excel, data_only=False)

    if nombre_hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo.")

    hoja = wb[nombre_hoja]
    precedentes = []  # Lista de tuplas (hoja, celda)
    celda = "E123"
    for row in hoja.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Si la celda contiene fórmula
                formula = cell.value

                # Buscar referencias como Hoja!A1 o simplemente A1
                referencias = re.findall(
                    r"(?:'([^']+)'|([A-Za-z0-9_]+))!([A-Z]{1,3}\d+)|([A-Z]{1,3}\d+)",
                    formula
                )

                for ref in referencias:
                    if ref[2]:  # Con nombre de hoja
                        hoja_ref = ref[0] or ref[1]
                        celda_ref = ref[2]
                    else:       # Mismo hoja
                        hoja_ref = nombre_hoja
                        celda_ref = ref[3]

                    precedentes.append((hoja_ref, celda_ref))

    return precedentes

def obtener_precedente_por_hoja_especifica(ruta_archivo, nombre_hoja):
    wb = openpyxl.load_workbook(ruta_archivo, data_only=False)
    sheet = wb[nombre_hoja]
    precedentes = []

    # Buscar fórmulas en celdas
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                # Extraer referencias con regex (como A1, B2, etc.)
                referencias = re.findall(r'([A-Z]+[0-9]+)', cell.value)
                precedentes.append({
                    "celda": cell.coordinate,
                    "formula": cell.value,
                    "referencias": referencias
                })
    return precedentes

def obtener_precedentes_con_valor(ruta, hoja_nombre):
    # Cargar dos veces: una con fórmula, otra con valores calculados
    wb_formulas = openpyxl.load_workbook(ruta, data_only=False)
    wb_valores = openpyxl.load_workbook(ruta, data_only=True)

    if hoja_nombre not in wb_formulas.sheetnames:
        return []

    hoja_formulas = wb_formulas[hoja_nombre]
    hoja_valores = wb_valores[hoja_nombre]

    precedentes = []
    formula_re = re.compile(r'([A-Za-z_]+!)?[A-Z]+\d+')  # también detecta referencias a otras hojas

    for row in hoja_formulas.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                celda_coord = cell.coordinate
                formula = cell.value
                valor = hoja_valores[cell.coordinate].value
    
                # Encontrar referencias en la fórmula
                #referencias = formula_re.findall(formula)
                #referencias = [ref.strip('!') for ref in referencias]
                referencias = re.findall(r'([A-Z]+[0-9]+)', cell.value)
                
                precedentes.append({
                    "celda": celda_coord,
                    "formula": formula,
                    "referencias": referencias,
                    "valor": valor
                })

    return precedentes

def unir_valores_numericos(valor):
    """
    Convierte un valor como '5,5,5,.,5,5' o ['5','5','5','.','5','5'] en '555.55'
    """
    if isinstance(valor, list):
        valor_str = ''.join(valor)
    elif isinstance(valor, str):
        valor_str = valor.replace(',', '')  # Eliminar comas
    else:
        return valor  # Retornar tal cual si no es str ni lista

    # Validar que el string resultante represente un número válido
    try:
        return float(valor_str)
    except ValueError:
        return valor_str

def obtener_precedentes_de_celda(ruta, hoja_nombre, celda_objetivo):


    # Cargar el archivo con las fórmulas visibles
    wb = openpyxl.load_workbook(ruta, data_only=False)

    if hoja_nombre not in wb.sheetnames:
        raise ValueError(f"La hoja '{hoja_nombre}' no existe en el archivo.")

    hoja = wb[hoja_nombre]
    celda = hoja[celda_objetivo]

    if not isinstance(celda.value, str) or not celda.value.startswith('='):
        return {
            "celda": celda_objetivo,
            "es_formula": False,
            "mensaje": "La celda no contiene una fórmula."
        }

    formula = celda.value

    # Expresión regular para referencias cruzadas: Hoja!A1 o simplemente A1
    formula_re = re.compile(r"(?:'[^']+'|[A-Za-z0-9_]+)?!?[A-Z]+\d+")
    referencias = formula_re.findall(formula)

    precedentes = []
    for ref in referencias:
        if '!' in ref:
            hoja_ref, celda_ref = ref.split('!')
            hoja_ref = hoja_ref.strip("'")  # quitar comillas si las hay
        else:
            hoja_ref = hoja_nombre
            celda_ref = ref

        try:
            valor = wb[hoja_ref][celda_ref].value
        except KeyError:
            valor = None

        precedentes.append({
            "hoja": hoja_ref,
            "celda": celda_ref,
            "valor": valor
        })

    return {
        "celda": celda_objetivo,
        "formula": formula,
        "es_formula": True,
        "precedentes": precedentes
    }

def obtener_precedentes_de_celda_con_valor(ruta, hoja_nombre, celda_objetivo):
    # Abrimos dos veces el archivo: uno con fórmulas, otro con valores ya evaluados
    wb_formulas = openpyxl.load_workbook(ruta, data_only=False)
    wb_valores = openpyxl.load_workbook(ruta, data_only=True)

    if hoja_nombre not in wb_formulas.sheetnames:
        raise ValueError(f"La hoja '{hoja_nombre}' no existe en el archivo.")

    hoja_formula = wb_formulas[hoja_nombre]
    hoja_valor = wb_valores[hoja_nombre]

    celda_formula = hoja_formula[celda_objetivo]
    celda_valor = hoja_valor[celda_objetivo]

    if not isinstance(celda_formula.value, str) or not celda_formula.value.startswith('='):
        return {
            "celda": celda_objetivo,
            "es_formula": False,
            "mensaje": "La celda no contiene una fórmula.",
            "valor": celda_valor.value,
            "precedentes" : [],
            "hoja":hoja_valor.title
        }

    formula = celda_formula.value

    # Buscar referencias en la fórmula
    formula_re = re.compile(r"(?:'[^']+'|[A-Za-z0-9_]+)?!?[A-Z]+\d+")
    referencias_raw = formula_re.findall(formula)

    precedentes = []
    for ref in referencias_raw:
        if '!' in ref:
            hoja_ref, celda_ref = ref.split('!')
            hoja_ref = hoja_ref.strip("'")
        else:
            hoja_ref = hoja_nombre
            celda_ref = ref

        try:
            valor_celda = wb_valores[hoja_ref][celda_ref].value
        except KeyError:
            valor_celda = None

        precedentes.append({
            "hoja": hoja_ref,
            "celda": celda_ref,
            "valor": valor_celda
        })

    return {
        "celda": celda_objetivo,
        "formula": formula,
        "valor_actual": celda_valor.value,
        "es_formula": True,
        "precedentes": precedentes
    }