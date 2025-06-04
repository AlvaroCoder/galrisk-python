from funciones import obtener_precedentes_de_celda_con_valor
import os
import openpyxl
import copy
import tempfile

def obtener_precedentes_recursivos(ruta, hoja_inicial, celda_inicial):

    visitados = set()  # Para evitar ciclos
    resultados = []

    # Cola de celdas a procesar (cada item es tupla: (hoja, celda))
    cola = [(hoja_inicial, celda_inicial)]

    while cola:
        hoja_actual, celda_actual = cola.pop(0)
        clave = f"{hoja_actual}!{celda_actual}"
        if clave in visitados:
            continue

        visitados.add(clave)

        # Llamamos a la función previa para obtener info de la celda
        resultado = obtener_precedentes_de_celda_con_valor(ruta, hoja_actual, celda_actual)
        resultados.append(resultado)

        if resultado["es_formula"]:
            for precedente in resultado["precedentes"]:
                hoja_pre = precedente["hoja"]
                celda_pre = precedente["celda"]
                clave_pre = f"{hoja_pre}!{celda_pre}"
                if clave_pre not in visitados:
                    cola.append((hoja_pre, celda_pre))

    return resultados

import openpyxl
import re

def obtener_precedentes_celda(ruta, hoja_nombre, celda_coord):
    wb_formulas = openpyxl.load_workbook(ruta, data_only=False)
    wb_valores = openpyxl.load_workbook(ruta, data_only=True)

    hoja_f = wb_formulas[hoja_nombre]
    hoja_v = wb_valores[hoja_nombre]

    cell_f = hoja_f[celda_coord]
    cell_v = hoja_v[celda_coord]

    if not isinstance(cell_f.value, str) or not cell_f.value.startswith('='):
        return [{
            "hoja": hoja_nombre,
            "celda": celda_coord,
            "es_formula": False,
            "mensaje": "La celda no contiene una fórmula.",
            "valor": cell_v.value
        }]

    formula = cell_f.value
    valor = cell_v.value

    # Detectar referencias como A1, B2, Hoja2!A3, etc.
    formula_re = re.compile(r'([A-Za-z_0-9]+!)?[A-Z]+\d+')
    coincidencias = formula_re.findall(formula)

    referencias = []
    for match in coincidencias:
        ref = match[0]
        if '!' in ref:
            hoja_ref, celda_ref = ref.split('!')
            hoja_ref = hoja_ref.strip()
            celda_ref = celda_ref.strip()
        else:
            hoja_ref = hoja_nombre
            celda_ref = ref.strip()

        referencias.append({"hoja": hoja_ref, "celda": celda_ref})

    return [{
        "hoja": hoja_nombre,
        "celda": celda_coord,
        "es_formula": True,
        "formula": formula,
        "valor": valor,
        "precedentes": referencias
    }]

def filtrar_raices_precedentes(lista_precedentes):

    raices = []
    for item in lista_precedentes:
        if not item.get("es_formula", True) and item.get("valor") is not None:
            raices.append({
                "hoja": item.get("hoja"),
                "celda": item.get("celda"),
                "valor": item.get("valor")
            })
    return raices

def simular_impacto_precedentes(ruta, hoja_objetivo, celda_objetivo, raices):
    """
    Simula el impacto en la celda objetivo cuando se incrementa en un 1% el valor de cada raíz.

    Args:
        ruta (str): Ruta del archivo Excel.
        hoja_objetivo (str): Nombre de la hoja donde está la celda objetivo.
        celda_objetivo (str): Coordenada de la celda objetivo (ej. 'H10').
        raices (list): Lista de diccionarios con raíces (hoja, celda, valor).

    Returns:
        list: Resultados con hoja, celda raíz, valor original, valor nuevo, valor objetivo original, valor objetivo nuevo, y variación.
    """
    resultados = []

    # Cargar el archivo Excel con valores reales
    wb_original = openpyxl.load_workbook(ruta, data_only=True)
    if hoja_objetivo not in wb_original.sheetnames:
        raise ValueError(f"La hoja {hoja_objetivo} no existe en el archivo.")

    valor_objetivo_original = wb_original[hoja_objetivo][celda_objetivo].value
    for raiz in raices:
        hoja_raiz = raiz["hoja"]
        celda_raiz = raiz["celda"]
        valor_raiz = raiz["valor"]
        if hoja_raiz not in wb_original.sheetnames:
            continue  # Omitir si la hoja no existe

        if not isinstance(valor_raiz, (int, float)):
            continue  # Solo operar sobre valores numéricos

        # Crear copia del archivo para no afectar el original
        wb_modificado = openpyxl.load_workbook(ruta, data_only=False, keep_vba=False)
        hoja_mod = wb_modificado[hoja_raiz]

        # Modificar valor raíz en +1%
        nuevo_valor = valor_raiz * 1.01
        hoja_mod[celda_raiz].value = nuevo_valor

        # Guardar en archivo temporal
        
        ruta_temp = "/Users/alvarofelipepupuchemorales/Desktop/Proyecto Economia/temp/temp_simulacion.xlsx"
        wb_modificado.save(ruta_temp)

        # Cargar nuevamente para obtener nuevo valor objetivo
        wb_ejecutado = openpyxl.load_workbook(ruta_temp, data_only=True)
        nuevo_valor_objetivo = wb_ejecutado[hoja_objetivo][celda_objetivo].value
        print(nuevo_valor_objetivo)
        # Calcular variación
        variacion = None
        if isinstance(nuevo_valor_objetivo, (int, float)) and isinstance(valor_objetivo_original, (int, float)):
            variacion = nuevo_valor_objetivo - valor_objetivo_original
        print(hoja_raiz, celda_raiz, )
        resultados.append({
            "hoja_raiz": hoja_raiz,
            "celda_raiz": celda_raiz,
            "valor_raiz_original": valor_raiz,
            "valor_raiz_modificado": nuevo_valor,
            "valor_objetivo_original": valor_objetivo_original,
            "valor_objetivo_nuevo": nuevo_valor_objetivo,
            "variacion": variacion,
            "variacion_absoluta" : abs(variacion)
        })
        # Limpiar archivo temporal
        if os.path.exists(ruta_temp):
            os.remove(ruta_temp)

    return resultados

from tabulate import tabulate

def lista_resultados_influyentes(resultados):
    return sorted(resultados, key=lambda x:x.get("variacion_absoluta", 0), reverse=True)

def mostrar_resultados_tabla(resultados):
    """
    Muestra los resultados de la simulación en formato de tabla.

    Args:
        resultados (list): Lista de diccionarios con los resultados de la simulación.
    """
    if not resultados:
        print("No hay resultados para mostrar.")
        return

    tabla = []

    resultados_ordenados = sorted(resultados, key=lambda x:x.get("variacion_absoluta", 0), reverse=True)

    for r in resultados_ordenados[:5]:
        fila = [
            r.get("hoja_raiz"),
            r.get("celda_raiz"),
            round(r.get("valor_raiz_original", 0), 6),
            round(r.get("valor_raiz_modificado", 0), 6),
            round(r.get("valor_objetivo_original", 0), 6),
            r.get("valor_objetivo_nuevo", 0),
            round(r.get("variacion", 0), 6) if r.get("variacion") is not None else "N/A",
            r.get("variacion_absoluta")
        ]
        tabla.append(fila)

    headers = [
        "Hoja Raíz", "Celda Raíz", "Valor Original", "Valor Modificado (+1%)",
        "Valor Objetivo Original", "Valor Objetivo Nuevo", "Variación", "Variación Absoluta"
    ]

    print(tabulate(tabla, headers=headers, tablefmt="grid"))

import openpyxl
import tempfile
import os
import shutil
import subprocess
import time

def simular_impacto_precedentes_2(ruta="", hoja_objetivo="", celda_objetivo="", raices=[]):

    """
    Simula el impacto en la celda objetivo cuando se incrementa en un 1% el valor de cada raíz.

    Args:
        ruta (str): Ruta del archivo Excel.
        hoja_objetivo (str): Nombre de la hoja donde está la celda objetivo.
        celda_objetivo (str): Coordenada de la celda objetivo (ej. 'H10').
        raices (list): Lista de diccionarios con raíces (hoja, celda, valor).

    Returns:
        list: Resultados con hoja, celda raíz, valor original, valor nuevo, valor objetivo original, valor objetivo nuevo, y variación.
    """
    resultados = []

    # Leer valor objetivo original
    wb_original = openpyxl.load_workbook(ruta, data_only=True)
    if hoja_objetivo not in wb_original.sheetnames:
        raise ValueError(f"La hoja {hoja_objetivo} no existe en el archivo.")
    valor_objetivo_original = wb_original[hoja_objetivo][celda_objetivo].value

    for raiz in raices:
        hoja_raiz = raiz["hoja"]
        celda_raiz = raiz["celda"]
        valor_raiz = raiz["valor"]

        if not isinstance(valor_raiz, (int, float)):
            continue
        if hoja_raiz not in wb_original.sheetnames:
            continue

        # Modificar el archivo con el nuevo valor
        wb_modificado = openpyxl.load_workbook(ruta, data_only=False)
        hoja_mod = wb_modificado[hoja_raiz]
        nuevo_valor = valor_raiz * 1.01
        hoja_mod[celda_raiz].value = nuevo_valor

        # Crear archivo temporal
        temp_dir = tempfile.mkdtemp()
        ruta_temp = os.path.join(temp_dir, "temp.xlsx")
        wb_modificado.save(ruta_temp)

        # Forzar apertura con Excel para recalcular fórmulas
        try:
            subprocess.run([
                "osascript", "-e",
                f'''
                tell application "Microsoft Excel"
                    activate
                    set wb to open POSIX file "{ruta_temp}"
                    delay 2
                    calculate wb
                    save wb
                    close wb saving yes

                    if (count of workbooks) = 0 then
                        quit saving yes
                    end if
                end tell
                '''
            ])
            time.sleep(5)
        except Exception as e:
            print(f"Error abriendo Excel: {e}")

        # Volver a abrir archivo con fórmulas actualizadas
        wb_ejecutado = openpyxl.load_workbook(ruta_temp, data_only=True)
        nuevo_valor_objetivo = wb_ejecutado[hoja_objetivo][celda_objetivo].value
        variacion = None
        if isinstance(nuevo_valor_objetivo, (int, float)) and isinstance(valor_objetivo_original, (int, float)):
            variacion = nuevo_valor_objetivo - valor_objetivo_original
        
        print(hoja_raiz, celda_raiz, nuevo_valor_objetivo, variacion, abs(variacion))
        resultados.append({
            "hoja_raiz": hoja_raiz,
            "celda_raiz": celda_raiz,
            "valor_raiz_original": valor_raiz,
            "valor_raiz_modificado": nuevo_valor,
            "valor_objetivo_original": valor_objetivo_original,
            "valor_objetivo_nuevo": nuevo_valor_objetivo,
            "variacion": variacion,
            "variacion_absoluta" : abs(variacion)
        })

        # Limpiar archivos temporales
        wb_ejecutado.close()
        shutil.rmtree(temp_dir)

    return resultados