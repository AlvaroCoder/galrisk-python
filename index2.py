


ruta_excel = "/Users/alvarofelipepupuchemorales/Desktop/Proyecto Economia/assets/PALTA HASS  BCP.xlsx"
hoja_nombre = "Costos Agricolas 01 ha"
celda_objetivo = "E123"

import openpyxl
import re

def generar_precedentes(ruta_excel, hoja_objetivo, celda_objetivo):
    wb = openpyxl.load_workbook(ruta_excel, data_only=False)
    hoja = wb[hoja_objetivo]

    mapa = {}
    formulas = {}
    raices = {}
    visitadas = set()

    def npv(rate, values):
        return sum(v / (1 + rate) ** (i + 1) for i, v in enumerate(values))

    def obtener_formula(celda):
        valor = hoja[celda].value
        if hoja[celda].data_type == 'f':
            return valor
        return None

    def limpiar_formula(formula):
        if formula.startswith('='):
            formula = formula[1:]
        if formula.startswith('+'):
            formula = formula[1:]
        return formula

    def extraer_referencias(formula):
        sin_hoja = re.sub(r"(?:'[^']+'|[A-Za-z0-9_]+)!", "", formula)
        return re.findall(r'\b[A-Z]{1,3}[0-9]{1,7}\b', sin_hoja)

    def expandir_rango(rango):
        from openpyxl.utils import column_index_from_string, get_column_letter
        start, end = rango.split(":")
        start_col, start_row = re.match(r'([A-Z]+)([0-9]+)', start).groups()
        end_col, end_row = re.match(r'([A-Z]+)([0-9]+)', end).groups()

        cols = list(range(column_index_from_string(start_col), column_index_from_string(end_col) + 1))
        rows = list(range(int(start_row), int(end_row) + 1))

        resultado = [f"{get_column_letter(col)}{row}" for col in cols for row in rows]
        return resultado

    def procesar_celda(celda):
        if celda in visitadas:
            return
        visitadas.add(celda)

        formula = obtener_formula(celda)
        if not formula:
            raices[celda] = hoja[celda].value
            return

        formula_limpia = limpiar_formula(formula)
        formula_sin_hoja = re.sub(r"(?:'[^']+'|[A-Za-z0-9_]+)!", "", formula_limpia)

        # Expandir rangos y usar placeholders temporales
        rangos = re.findall(r'\$?[A-Z]{1,3}\$?[0-9]+:\$?[A-Z]{1,3}\$?[0-9]+', formula_sin_hoja)
        placeholders = {}
        for i, rango in enumerate(rangos):
            celdas = expandir_rango(rango.replace('$', ''))
            reemplazo = "[" + ",".join(f"r('{c}')" for c in celdas) + "]"
            placeholder = f"__RANGO_{i}__"
            formula_sin_hoja = formula_sin_hoja.replace(rango, placeholder)
            placeholders[placeholder] = reemplazo

        referencias = extraer_referencias(formula_sin_hoja)
        referencias = [ref.replace('$', '') for ref in referencias]

        for ref in referencias:
            procesar_celda(ref)

        # Reemplazar referencias simples
        for ref in referencias:
            formula_sin_hoja = re.sub(rf'\b{ref}\b', f"r('{ref}')", formula_sin_hoja)

        # Restaurar los rangos
        for placeholder, reemplazo in placeholders.items():
            formula_sin_hoja = formula_sin_hoja.replace(placeholder, reemplazo)

        formulas[celda] = formula_sin_hoja

        try:
            mapa[celda] = eval(f"lambda r: {formula_sin_hoja}", {
                "npv": npv,
                "sum": sum,
                "max": max,
                "min": min,
                "abs": abs
            })
        except Exception as e:
            print(f"[ERROR] al evaluar fórmula en {celda}: {formula_sin_hoja}")
            print(f"       {e}")

    procesar_celda(celda_objetivo)
    return mapa, formulas, raices
mapa, formulas, raices = generar_precedentes(ruta_excel, hoja_nombre, celda_objetivo)


