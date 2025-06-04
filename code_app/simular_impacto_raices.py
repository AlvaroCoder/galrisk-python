import openpyxl
import tempfile
import os
import shutil
import subprocess
import time

def simular_impacto_raices(ruta, hoja_objetivo, celda_objetivo, raices_dict):
    """
    Simula el impacto en la celda objetivo cuando se incrementa en un 1% el valor de cada raíz.

    Args:
        ruta (str): Ruta del archivo Excel.
        hoja_objetivo (str): Nombre de la hoja donde está la celda objetivo.
        celda_objetivo (str): Coordenada de la celda objetivo (ej. 'H10').
        raices_dict (dict): Diccionario de raíces como {"Hoja!Celda": valor}

    Returns:
        list: Resultados en forma de diccionario.
    """
    resultados = []

    # Leer valor objetivo original
    wb_original = openpyxl.load_workbook(ruta, data_only=True)
    if hoja_objetivo not in wb_original.sheetnames:
        raise ValueError(f"La hoja {hoja_objetivo} no existe en el archivo.")
    valor_objetivo_original = wb_original[hoja_objetivo][celda_objetivo].value

    for referencia, valor_raiz in raices_dict.items():
        hoja_raiz, celda_raiz = referencia.split("!")
        if not isinstance(valor_raiz, (int, float)):
            continue
        if hoja_raiz not in wb_original.sheetnames:
            continue

        # Modificar el archivo con el nuevo valor
        wb_modificado = openpyxl.load_workbook(ruta, data_only=False)
        hoja_mod = wb_modificado[hoja_raiz]
        nuevo_valor = valor_raiz * 1.01
        hoja_mod[celda_raiz].value = nuevo_valor

        # Crear archivo temporal
        temp_dir = tempfile.mkdtemp()
        ruta_temp = os.path.join(temp_dir, "temp.xlsx")
        wb_modificado.save(ruta_temp)

        # Forzar recálculo en Excel
        try:
            subprocess.run([
                "osascript", "-e",
                f'''
                tell application "Microsoft Excel"
                    activate
                    set wb to open POSIX file "{ruta_temp}"
                    delay 2
                    calculate wb
                    save wb
                    close wb saving yes
                    if (count of workbooks) = 0 then quit saving yes
                end tell
                '''
            ])
            time.sleep(5)
        except Exception as e:
            print(f"Error al recalcular con Excel: {e}")

        # Leer resultado actualizado
        wb_ejecutado = openpyxl.load_workbook(ruta_temp, data_only=True)
        nuevo_valor_objetivo = wb_ejecutado[hoja_objetivo][celda_objetivo].value
        variacion = None
        if isinstance(nuevo_valor_objetivo, (int, float)) and isinstance(valor_objetivo_original, (int, float)):
            variacion = nuevo_valor_objetivo - valor_objetivo_original

        resultados.append({
            "Hoja Raíz": hoja_raiz,
            "Celda Raíz": celda_raiz,
            "Valor Original": valor_raiz,
            "Valor Modificado": nuevo_valor,
            "Valor Objetivo Original": valor_objetivo_original,
            "Valor Objetivo Nuevo": nuevo_valor_objetivo,
            "Variación": variacion,
            "Variación Absoluta": abs(variacion) if variacion is not None else None
        })

        # Limpieza
        wb_ejecutado.close()
        shutil.rmtree(temp_dir)

    return resultados


def calcular_npv(tasa, flujos):
    return sum(f / (1 + tasa) ** t for t, f in enumerate(flujos, start=1))

def simular_impacto_sin_excel(raices, nombre_tasa, nombres_flujos, variacion=0.05):
    """
    Simula impacto modificando las raíces sin abrir el archivo Excel.

    - raices: dict con todas las celdas clave y sus valores
    - nombre_tasa: clave que representa la tasa de descuento (ej. 'A1')
    - nombres_flujos: lista con las claves de los flujos de caja (ej. ['B1', 'B2', 'B3'])
    """
    tasa_base = raices[nombre_tasa]
    flujos_base = [raices[n] for n in nombres_flujos]

    # Valor base del NPV
    npv_base = calcular_npv(tasa_base, flujos_base)

    impactos = []

    for nombre in [nombre_tasa] + nombres_flujos:
        valor_original = raices[nombre]
        
        # Modificar ± variación%
        nuevos_valores = [
            valor_original * (1 + variacion),
            valor_original * (1 - variacion)
        ]

        for nuevo_valor in nuevos_valores:
            raices_mod = raices.copy()
            raices_mod[nombre] = nuevo_valor

            nueva_tasa = raices_mod[nombre_tasa]
            nuevos_flujos = [raices_mod[n] for n in nombres_flujos]
            nuevo_npv = calcular_npv(nueva_tasa, nuevos_flujos)

            variacion_abs = abs(nuevo_npv - npv_base)
            impactos.append({
                'celda': nombre,
                'valor_original': valor_original,
                'valor_modificado': nuevo_valor,
                'npv_resultante': nuevo_npv,
                'variacion_absoluta': variacion_abs
            })

    # Ordenar por mayor impacto
    impactos_ordenados = sorted(impactos, key=lambda x: x['variacion_absoluta'], reverse=True)

    return npv_base, impactos_ordenados[:6]

